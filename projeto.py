import tkinter as tk
from tkinter import ttk, messagebox
from tkinter import PhotoImage
from PIL import Image, ImageTk
import openpyxl

#centralização de janelas
def center_window(window):
    window.update_idletasks()
    width = window.winfo_width()
    height = window.winfo_height()
    x = (window.winfo_screenwidth() // 2) - (width // 2)
    y = (window.winfo_screenheight() // 2) - (height // 2)
    window.geometry('{}x{}+{}+{}'.format(width, height, x, y))

#função para limpar a janela main
def clear_main():
    # Identifica entrys, labels ou frames no layout da janela
    for widget in janela_main.winfo_children():
        if isinstance(widget, (tk.Entry, tk.Label, tk.Button)) and widget not in [cadastro_fornecedores_button, cadastro_auditorias_button]:
            widget.destroy()
        elif isinstance(widget, tk.Frame):
            # Destruir todos os widgets dentro do frame
            for frame_widget in widget.winfo_children():
                frame_widget.destroy()
            # Destruir o frame em si
            widget.destroy()

#função para fechar todas as janelas do programa
def fechar_janela(janela):
    janela.destroy()

#função para reiniciar o programa após finalizar registro de usuário
def reiniciar_programa():
    if 'janela_main' in globals():
        janela_main.destroy()
    if 'janela_registro' in globals():
        janela_registro.destroy()
    if 'janela_login' in globals():
        janela_login.destroy()

    # Chamar a função login para reiniciar o programa
    login()

#Mensagem de ação concluída (auditoria ou fornecedor)
def show_custom_message(janela_main):
    custom_message = tk.Toplevel(janela_main)
    custom_message.title("Ação Concluída!")
    custom_message.geometry("300x150")

    label_info = tk.Label(custom_message, text="Dados salvos em Excel! Reinicie ou Faça novo Registro.")
    label_info.pack(pady=10)

    restart_button = tk.Button(custom_message, text="Reinciar Programa", command=reiniciar_programa)
    restart_button.pack(side=tk.RIGHT, padx=10, pady=10)

    ok_button = tk.Button(custom_message, text="OK", command=lambda: fechar_janela(custom_message))
    ok_button.pack(side=tk.LEFT, padx=10, pady=10)

#função de validação de numeros no campo data
def validade_input(input_str):
    if input_str.isdigit() or input_str =="":
        return True
    else: 
        return False

#configuração de janela_main após clicar no botão de "registrar auditoria"
def main_config_audit():
    clear_main()
    global data_audit_entry
    janela_main.title("Cadastro de Auditorias")

    def export_audit_excel():
        #abrir arquivo existente
        planilha_sup = openpyxl.load_workbook("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Auditorias\\Auditorias.xlsx")

        #abrindo a primeira planilha ativa
        sheet_sup = planilha_sup.active

        #coletar os dados inseridos pelo usuário
        supply_data = [data_audit_entry.get(), rej_int_entry.get(), specprod_entry.get(), armazenamento_entry.get(), spec_test_entry.get(), field_entry.get(), postos_entry.get()]

        #encontrar a proxima linha vazia
        next_row_colab = sheet_sup.max_row + 1

        #função para coletar os dados em supply_data e dar a eles indice começando em 1
        for col, valor in enumerate(supply_data, start=1):
            sheet_sup.cell(row=next_row_colab, column = col, value = valor)

        #salvar a planilha após registro
        planilha_sup.save("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Auditorias\\Auditorias.xlsx")
        #mensagem de ação completa 
        show_custom_message(janela_main)

    #botão de registro de auditoria
    audit_register_button = tk.Button(janela_main, text="Registro de Auditoria", command=export_audit_excel)
    audit_register_button.grid(row=11, column=1, columnspan=2, padx=10, pady=10)

    # Frame para o cabeçalho
    frame_cabecalho = tk.Frame(janela_main)
    frame_cabecalho.grid(row=0, column=1, columnspan=3, padx=10, pady=10)

    # Texto do cabeçalho
    texto_informativo = tk.Label(frame_cabecalho, text="DECATHLON: Movendo pessoas através das maravilhas do esporte com vitalidade, responsabilidade, generosidade e autenticidade")
    texto_informativo.pack()

    # Frame para a imagem
    frame_imagem = tk.Frame(janela_main)
    frame_imagem.grid(row=1, column=1, padx=(180, 0), pady=10)

    # Carregar a imagem do logotipo da Decathlon
    img_logo = Image.open("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Imagens\\Decathlon.png")
    img_logo = img_logo.resize((220, 30))
    img_logo = ImageTk.PhotoImage(img_logo)

    # Exibir a imagem
    img_logo_label = tk.Label(frame_imagem, image=img_logo)
    img_logo_label.image = img_logo
    img_logo_label.pack()
    
    #data
    data_audit_label = tk.Label(janela_main, text="Data:")
    data_audit_label.grid(row=4, column=1, padx=10, pady=10, sticky="e")

    #vincula o entry a u função date_input_barra_auto
    data_audit_var = tk.StringVar()

    #armazena len do entry, compara com 2 e 5 para adicionar / quando preciso
    def on_date_entry_change(*args):
        current_data_text = data_audit_var.get()
        if len(current_data_text) == 2 or len(current_data_text) == 5:
            data_audit_var.set(current_data_text + "/")
            data_audit_entry.after(10, move_cursor)
    def move_cursor():
        data_audit_entry.icursor(tk.END)

    data_audit_entry = tk.Entry(janela_main, width=25, validate="key", validatecommand=(validar_data_input_cmd, "%P"), textvariable=data_audit_var)
    data_audit_entry.grid(row=4, column=2, padx=10, pady=10)

    #vincular a funçção a mudançã de caracterez armazenada current_data_text
    data_audit_var.trace_add('write', on_date_entry_change)

    #rej int (%)
    rej_int_label = tk.Label(janela_main, text="Rej. Int. (%):")
    rej_int_label.grid(row=5, column=1, padx=10, pady=10, sticky="e")
    rej_int_entry = tk.Entry(janela_main, width=25)
    rej_int_entry.grid(row=5, column=2, padx=10, pady=10)
    #Spec. Produto
    specprod_label = tk.Label(janela_main, text="Spec. Produto:")
    specprod_label.grid(row=6, column=1, padx=10, pady=10, sticky="e")
    specprod_combobox = ttk.Combobox(janela_main, values=["OK", "NG"], state="readonly", width=22)
    specprod_combobox.grid(row=6, column=2, padx=10, pady=10)
    #Spec. Armazenamento
    armazenamento_label = tk.Label(janela_main, text="Spec. Armazenamento:")
    armazenamento_label.grid(row=7, column=1, padx=10, pady=10, sticky="e")
    specarm_combobox = ttk.Combobox(janela_main, values=["OK", "NG"], state="readonly", width=22)
    specarm_combobox.grid(row=7, column=2, padx=10, pady=10)
    #Spec. Testes
    spectest_label = tk.Label(janela_main, text="Spec. Testes:")
    spectest_label.grid(row=8, column=1, padx=10, pady=10, sticky="e")
    spectest_combobox = ttk.Combobox(janela_main, values=["OK", "NG"], state="readonly", width=22)
    spectest_combobox.grid(row=8, column=2, padx=10, pady=10)
    #field (%)
    field_label = tk.Label(janela_main, text="Lotes Reprovados (%):")
    field_label.grid(row=9, column=1, padx=10, pady=10, sticky="e")
    field_entry = tk.Entry(janela_main, width=25)
    field_entry.grid(row=9, column=2, padx=10, pady=10)

    #para o campo fornecedor, utilizaremos os dados da planilha fornecedores
    def fornecedores_planilha():
        fornecedores_ws = openpyxl.load_workbook("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Fornecedores\\Fornecedores.xlsx")
        open_fornecedores_ws = fornecedores_ws.active
        fornecedores_list = []

        #iteração das linhas da planilha para extrair os dados
        for row in open_fornecedores_ws.iter_rows(min_row=2, values_only=True):
            #coletar apenas o nome do fornecedor (coluna A)
            fornecedor = row[0]
            if fornecedor:
                fornecedores_list.append(fornecedor)
        return fornecedores_list

    #chamando a função para obter os dados de fornecedores da planilha
    fornecedores_data_loaded = fornecedores_planilha()

    #Fornecedor (audit)
    supply_audit_label = tk.Label(janela_main, text="Fornecedor:")
    supply_audit_label.grid(row=10, column=1, padx=10, pady=10, sticky="e")
    supply_combobox = ttk.Combobox(janela_main, values=fornecedores_data_loaded, state="readonly", width=22)
    supply_combobox.grid(row=10, column=2, padx=10, pady=10)

    # Definindo uma função para fechar todas as janelas ao fechar o programa
    janela_main.protocol("WM_DELETE_WINDOW", fechar_programa)

#configuração de janela_main após clicar no botão de "registrar fornecedor"
def main_config_supply():
    clear_main()
    janela_main.title("Cadastro de fornecedores")

    def export_supply_excel():
        #abrir arquivo existente
        planilha_sup = openpyxl.load_workbook("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Fornecedores\\Fornecedores.xlsx")

        #abrindo a primeira planilha ativa
        sheet_sup = planilha_sup.active

        #coletar os dados inseridos pelo usuário
        supply_data = [nome_supply_entry.get(), razao_supply_entry.get(), produtos_supply_entry.get(), colab_supply_entry.get(), no_line_entry.get(), rua_no_entry.get(), bairro_entry.get()]

        #encontrar a proxima linha vazia
        next_row_colab = sheet_sup.max_row + 1

        #função para coletar os dados em supply_data e dar a eles indice começando em 1
        for col, valor in enumerate(supply_data, start=1):
            sheet_sup.cell(row=next_row_colab, column = col, value = valor)

        #salvar a planilha após registro
        planilha_sup.save("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Fornecedores\\Fornecedores.xlsx")

        #mensagem de ação completa 
        show_custom_message(janela_main)

    #botão de registro de fornecedor
    supply_register_button = tk.Button(janela_main, text="Registro de Fornecedor", command=export_supply_excel)
    supply_register_button.grid(row=11, column=1, columnspan=2, padx=10, pady=10)

    # Frame para o cabeçalho
    frame_cabecalho = tk.Frame(janela_main)
    frame_cabecalho.grid(row=0, column=1, columnspan=3, padx=10, pady=10)

    texto_informativo = tk.Label(frame_cabecalho, text="DECATHLON: Movendo pessoas através das maravilhas do esporte com vitalidade, responsabilidade, generosidade e autenticidade")
    texto_informativo.pack()

    #nome fornecedor
    nome_supply_label = tk.Label(janela_main, text="Nome do Fornecedor:")
    nome_supply_label.grid(row=4, column=1, padx=10, pady=10, sticky="e")
    nome_supply_entry = tk.Entry(janela_main, width=25)
    nome_supply_entry.grid(row=4, column=2, padx=10, pady=10)

    #razão social
    razao_supply_label = tk.Label(janela_main, text="Razão Social:")
    razao_supply_label.grid(row=5, column=1, padx=10, pady=10, sticky="e")
    razao_supply_entry = tk.Entry(janela_main, width=25)
    razao_supply_entry.grid(row=5, column=2, padx=10, pady=10)

    #produtos fabricados
    produtos_supply_label = tk.Label(janela_main, text="Produtos:")
    produtos_supply_label.grid(row=6, column=1, padx=10, pady=10, sticky="e")
    produtos_supply_entry = tk.Entry(janela_main, width=25)
    produtos_supply_entry.grid(row=6, column=2, padx=10, pady=10)

    #no. de colaboradores
    colab_supply_label = tk.Label(janela_main, text="No. Colaboradores:")
    colab_supply_label.grid(row=7, column=1, padx=10, pady=10, sticky="e")
    colab_supply_entry = tk.Entry(janela_main, width=25)
    colab_supply_entry.grid(row=7, column=2, padx=10, pady=10)

    #no de linhas
    no_line_label = tk.Label(janela_main, text="No. de Linhas:")
    no_line_label.grid(row=8, column=1, padx=10, pady=10, sticky="e")
    no_line_entry = tk.Entry(janela_main, width=25)
    no_line_entry.grid(row=8, column=2, padx=10, pady=10)

    #Rua, n.
    rua_no_label = tk.Label(janela_main, text="Rua, nº:")
    rua_no_label.grid(row=9, column=1, padx=10, pady=10, sticky="e")
    rua_no_entry = tk.Entry(janela_main, width=25)
    rua_no_entry.grid(row=9, column=2, padx=10, pady=10)

    #bairro
    bairro_label = tk.Label(janela_main, text="Bairro:")
    bairro_label.grid(row=10, column=1, padx=10, pady=10, sticky="e")
    bairro_entry = tk.Entry(janela_main, width=25)
    bairro_entry.grid(row=10, column=2, padx=10, pady=10)

    # Frame para a imagem
    frame_imagem = tk.Frame(janela_main)
    frame_imagem.grid(row=1, column=1, padx=(180, 0), pady=10)

    # Carregar a imagem do logotipo da Decathlon
    img_logo = Image.open("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Imagens\\Decathlon.png")
    img_logo = img_logo.resize((220, 30))
    img_logo = ImageTk.PhotoImage(img_logo)

    # Exibir a imagem
    img_logo_label = tk.Label(frame_imagem, image=img_logo)
    img_logo_label.image = img_logo
    img_logo_label.pack()

    janela_main.protocol("WM_DELETE_WINDOW", fechar_programa)

#confirguração de fechamento do programa
def fechar_programa():
    #destruir a janela main caso ainda exista após fechar a janela
    if 'janela_main' in globals():
        janela_main.destroy()
    if 'janela_registro' in globals():
        janela_registro.destroy()
    # Feche a janela de login
    janela_login.destroy()

#configuração de abertura da janela principal
def open_main():
    global janela_main, cadastro_auditorias_button, cadastro_fornecedores_button, validar_data_input_cmd
    janela_main = tk.Toplevel()
    janela_main.title("Cadastro de Auditorias e Fornecedores DECATHLON")

    #tamanho de janela
    janela_main.geometry("1000x500")

    #posicionamento da janela
    janela_main.resizable(False,False)
    center_window(janela_main)

    #formato digitação campo de data
    validar_data_input_cmd = janela_main.register(validade_input)

    #missão e valores empresa
    texto_informativo = tk.Label(janela_main, text="DECATHLON: Movendo pessoas através das maravilhas do esporte com vitalidade, responsabilidade, generosidade e autenticidade")
    texto_informativo.grid(row=0, column=1, columnspan=2, padx=10, pady=10)

    #imagem logo decathlon
    img_logo = Image.open("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Imagens\\Decathlon.png")
    img_logo = img_logo.resize((220, 30))
    img_logo = ImageTk.PhotoImage(img_logo)

    #posicionamento da imagem logo
    img_logo_label = tk.Label(janela_main, image=img_logo)
    img_logo_label.image = img_logo
    img_logo_label.grid(row=1, column=1, rowspan=2, padx=(180, 0), pady=10)

    #botões de cadastro
    #cadastro fornecedores
    cadastro_fornecedores_button = tk.Button(janela_main, text="Cadastro de fornecedores", width=30, command=main_config_supply)
    cadastro_fornecedores_button.grid(row=1, column=0, padx=10, pady=10)

    #cadastro auditorias
    cadastro_auditorias_button = tk.Button(janela_main, text="Cadastro de Auditorias", width=30, command=main_config_audit)
    cadastro_auditorias_button.grid(row=2, column=0, padx=10, pady=10)
    
    # Definindo uma função para fechar todas as janelas ao fechar o programa
    janela_main.protocol("WM_DELETE_WINDOW", fechar_programa)

#configuração da tela de login
def login():
    email = email_entry.get()
    senha = senha_entry.get()
    # Aqui você pode adicionar a lógica de autenticação
    print("E-mail:", email)
    print("Senha:", senha)
    janela_login.withdraw()  # Esconde a janela de login
    open_main()

#configuração da tela de registro de usuário
def open_register():
    global janela_registro
    janela_registro = tk.Toplevel()
    janela_registro.title("Registro de Colaborador")
    janela_registro.pack_propagate(False)

    #função para exportar os dados para excel
    def export_reg_excel():
        #abrindo o arquivo existente
        planilha_colab = openpyxl.load_workbook("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Colaboradores\\Colaboradores.xlsx")
        
        #abrindo a primeira planilha ativa
        sheet_colab = planilha_colab.active

        #coletar os dados inseridos pelo usuário
        colab_data = [nome_entry.get(), matricula_entry.get(), idade_entry.get(), unidade_entry.get(), funcao_combobox.get(), esporte_entry.get()]

        #encontra a proxima linha vazia na coluna A
        next_row = sheet_colab.max_row + 1

        #função para coletar os dados em colab_data e dar um índice para cada (começando de 1)
        for col, valor in enumerate(colab_data, start=1):
            sheet_colab.cell(row=next_row, column=col, value=valor)

        #salvar o arquivo após alteração
        planilha_colab.save("C:\\Users\\Gabriel Siza\\OneDrive\\Área de Trabalho\\Documents\\GABRIELIMPORTANTE\\Programing\\Python\\Aula 1\\Projeto Cadastro de Auditorias e Empreas [DECATHLON]\\Planilhas de Controle\\Registro de Colaboradores\\Colaboradores.xlsx")


    #botão de registro 
    register_excel_button = tk.Button(janela_registro, text="Registrar Colaborador", command=export_reg_excel)
    register_excel_button.grid(row=6, column=1, padx=10, pady=10)

    #titulos dos campos de input
    #nome
    nome_label = tk.Label(janela_registro, text="Nome: ")
    nome_label.grid(row=0, column=0, padx=10, pady=10, sticky="e")
    nome_entry = tk.Entry(janela_registro, width=28)
    nome_entry.grid(row=0, column=1, padx=10, pady=10)
    #matricula
    matricula_label = tk.Label(janela_registro, text="Matrícula: ")
    matricula_label.grid(row=1, column=0, padx=10, pady=10, sticky="e")
    matricula_entry = tk.Entry(janela_registro, width=28)
    matricula_entry.grid(row=1, column=1, padx=10, pady=10)
    #idade
    idade_label = tk.Label(janela_registro, text="Idade: ")
    idade_label.grid(row=2, column=0, padx=10, pady=10, sticky="e")
    idade_entry = tk.Entry(janela_registro, width=28)
    idade_entry.grid(row=2, column=1, padx=10, pady=10)
    #unidade
    unidade_label = tk.Label(janela_registro, text="Unidade: ")
    unidade_label.grid(row=3, column=0, padx=10, pady=10, sticky="e")
    unidade_entry = tk.Entry(janela_registro, width=28)
    unidade_entry.grid(row=3, column=1, padx=10, pady=10)
    #função
    funcao_label = tk.Label(janela_registro, text="Função: ")
    funcao_label.grid(row=4, column=0, padx=10, pady=10, sticky="e")
    funcao_combobox = ttk.Combobox(janela_registro, values=["supplier", "auditor", "inspetor", "analista", "engenheiro", "chefe", "supervisor", "gerente"], state="readonly", width=25)
    funcao_combobox.grid(row=4, column=1, padx=10, pady=10)
    #esporte
    esporte_label = tk.Label(janela_registro, text="Esporte: ")
    esporte_label.grid(row=5, column=0, padx=10, pady=10, sticky="e")
    esporte_entry = tk.Entry(janela_registro, width=28)
    esporte_entry.grid(row=5, column=1, padx=10, pady=10)

    janela_registro.protocol("WM_DELETE_WINDOW", fechar_programa) 
    janela_login.withdraw()  # Esconde a janela de login

#configuração do início do programa
def inciar_login():
    #definindo variáveis globais para acessar em outras funções
    global janela_login, email_entry, senha_entry

    # Criação da janela de login
    janela_login = tk.Tk()
    janela_login.title("Login")
    janela_login.geometry('270x120')
    center_window(janela_login)

    # Rótulo e campo de entrada para o e-mail
    email_label = tk.Label(janela_login, text="E-mail:")
    email_label.grid(row=0, column=0, padx=10, pady=5, sticky="e")
    email_entry = tk.Entry(janela_login, width=25)
    email_entry.grid(row=0, column=1, padx=10, pady=5)

    # Rótulo e campo de entrada para a senha
    senha_label = tk.Label(janela_login, text="Senha:")
    senha_label.grid(row=1, column=0, padx=10, pady=5, sticky="e")
    senha_entry = tk.Entry(janela_login, show="*", width=25)
    senha_entry.grid(row=1, column=1, padx=10, pady=5)

    # Botão de login
    login_button = tk.Button(janela_login, text="Login", command=login)
    login_button.grid(row=2, sticky="w", column=1, padx=20, pady=10)

    # Botão register
    register_button = tk.Button(janela_login, text="Register", command=open_register)
    register_button.grid(row=2, sticky="e", column=1, columnspan=10, padx=40, pady=10)

    # Iniciando o loop da interface gráfica
    janela_login.mainloop()
inciar_login()


#alterados campos de rej interna e field (audit)
#alterados campos de rua, nº e bairro (supply)
#formato digitação campo de data
#formato input de specs